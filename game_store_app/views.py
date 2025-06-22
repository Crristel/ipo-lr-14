from django.shortcuts import render
import json

from .models import Product,Category,Maker,Basket,Basket_element,Order,OrderItem
from .forms import Filter
from django.views.decorators.http import require_POST

from django.contrib.auth.decorators import login_required

from django.contrib.auth import login, authenticate
from django.shortcuts import render, redirect
from .forms import UserRegistrationForm, LoginForm

from django.contrib import messages
from django.urls import reverse

from openpyxl import Workbook
from django.http import HttpResponse
from io import BytesIO
from django.core.mail import send_mail
from django.conf import settings
from django.core.mail import EmailMessage
from openpyxl.styles import Font

def home_page(request):
    return render(request, 'home_page.html') #render возвращает шаблон html на запрос пользователя

def about_me(request):
    return render(request,'about_me.html')
def game_store(request):
    return render(request,'game_store.html')

def speciality(request):
    specialties = []  
    with open("data/dump.json", 'r', encoding='utf-8') as file:
        read_file = json.load(file)
        
    for specialty in read_file:
        if specialty.get("model") == "data.speciality":
            specialty_data = {
                "code": specialty["fields"].get("code"),
                "pk": specialty.get("pk"),
                "title": specialty["fields"].get("title"),
                "educational": specialty["fields"].get("c_type"),
            }
            specialties.append(specialty_data)  
    
    return render(request, 'speciality.html',{'specialties': specialties})

def speciality_found(request):
    code = request.GET.get('code')
    speciality = []  
    with open("data/dump.json", 'r', encoding='utf-8') as file:
        read_file = json.load(file)
        for specialty in read_file:
            if specialty.get("model") == "data.specialty":
                if specialty["fields"].get("code") == code:
                    specialty_data = {
                        "code": specialty["fields"].get("code"),
                        "pk": specialty.get("pk"),
                        "title": specialty["fields"].get("title"),
                        "educational": specialty["fields"].get("c_type"),
                    }
                    speciality.append(specialty_data)  

    return render(request, "speciality_found.html", {'speciality': speciality}) 

def speciality_id(request,id):
    speciality = []  
    with open("data/dump.json", 'r', encoding='utf-8') as file:
        read_file = json.load(file)
        for specialty in read_file:
            if specialty.get("model") == "data.specialty":
                if int(specialty.get("pk")) == int(id):
                    specialty_data = {
                        "code": specialty["fields"].get("code"),
                        "pk": specialty.get("pk"),
                        "title": specialty["fields"].get("title"),
                        "educational": specialty["fields"].get("c_type"),
                    }
                    speciality.append(specialty_data)  

    return render(request, "speciality_found.html", {'speciality': speciality})
    
def product_list(request):
    products = Product.objects.all()
    form = Filter(request.GET or None)
    if form.is_valid(): #Сохраняются в cleaned_data автоматически
        if form.cleaned_data['category']:
            products = products.filter(category__name__iexact=form.cleaned_data['category']) # Модификатор сравнения, который означает: i - case-insensitive (регистронезависимый) и exact - точное совпадение
        if form.cleaned_data['maker']:
            products = products.filter(maker__name__iexact=form.cleaned_data['maker'])
        if form.cleaned_data['search']:
            products = products.filter(name__icontains=form.cleaned_data['search']) # регистронезависимый поиск подстроки
    
    return render(request, 'shop/product_list.html', {'products': products, 'form': form})

def product_detail(request,id):
    try:
        product = Product.objects.get(id=id)
    except Product.DoesNotExist:
            messages.error(request, "Товар не найден.")
            return redirect(request.META.get('HTTP_REFERER', reverse('catalog'))) #reverse('catalog') вернёт строку '/catalog/'  request.META —словарь с метаданными запроса (заголовки, IP, User-Agent и т. д.).
   # HTTP_REFERER - URL, с которого пришёл пользователь
    return render(request, 'shop/product_detail.html', {'product': product})

@login_required
def add_to_cart(request, product_id):
    try:
        product = Product.objects.get(id=product_id)
    except Product.DoesNotExist or  UnboundLocalError :
        messages.error(request, "Товар не найден.")
        return redirect('cart_view')
    basket, created = Basket.objects.get_or_create(user=request.user) #Ищет запись в БД по указанным параметрам (user=request.user). Если запись найдена, возвращает её/или создает новую
    quantity = int(request.POST.get('quantity', 1))
    
    basket_element, created = Basket_element.objects.get_or_create(
        basket=basket,
        product=product,
        defaults={'quantity': quantity}
    )
    if not created:
        basket_element.quantity += quantity
        basket_element.save()
        messages.success(request, "Количество товара обновлено в корзине.")
    else:
        messages.success(request, "Товар добавлен в корзину.")
    return redirect('cart_view')   

@login_required
@require_POST
def update_cart(request, item_id):
    try:
        basket_element = Basket_element.objects.get(id=item_id, basket__user=request.user)
    except Basket_element.DoesNotExist:
        messages.error(request, "Не удалось обновить корзину")
        return redirect('cart_view')

    try:
        quantity = int(request.POST.get('quantity', 1))
    except (ValueError, TypeError):
        quantity = 1

    if quantity <= 0:
        basket_element.delete()
        messages.success(request, "Товар удалён из корзины")
    elif quantity <= basket_element.product.stok_quantity:
        basket_element.quantity = quantity
        basket_element.save()
        messages.success(request, "Количество обновлено")
    else:
        messages.error(request, "Недостаточно товара на складе")

    return redirect('cart_view')

@login_required
def remove_from_cart(request, item_id):
    try:
        basket_element = Basket_element.objects.get(id=item_id, basket__user=request.user)
        basket_element.delete()
        messages.success(request, "Товар удалён из корзины")
    except Basket_element.DoesNotExist:
        messages.error(request, "Товар не найден в корзине")
    return redirect('cart_view')

@login_required
def cart_view(request):
    basket, created = Basket.objects.get_or_create(user=request.user)
    basket_elements = basket.basket_elements.select_related('product').all() #позволяет Django автоматически получить все связанные объекты за один запорос ( пример с автором и книгами)
    total = sum(item.element_cost() for item in basket_elements)
    context = {
        'cart_items': basket_elements,
        'total': total,
    }
    return render(request, 'shop/cart.html', context)





def register(request):
    if request.method == 'POST':
        form = UserRegistrationForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user) #функция аутентификации, которая "входит" пользователя в систему, создавая сессию и делая его авторизованным для текущего запроса (request). позволяет исп декораторы @login_required
            return redirect('cart_view')
    else:
        form = UserRegistrationForm()
    return render(request, 'register.html', {'form': form})

def user_login(request):
    if request.method == 'POST':
        form = LoginForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password) #встроенная функция для проверки подлинности пользователя (валидации логина и пароля). 
            #Возвращает объект user, если данные верны, или None, если нет.
            if user is not None:
                login(request, user)
                return redirect('cart_view')
    else:
        form = LoginForm()
    return render(request, 'login.html', {'form': form})

@login_required
def checkout(request):
    try:
        # Получаем корзину пользователя
        cart = Basket.objects.get(user=request.user)
        cart_items = Basket_element.objects.filter(basket=cart)
        
        # Проверяем, что корзина не пуста
        if not cart_items.exists():
            return redirect('cart_view')
        
        # Рассчитываем общую сумму
        total_price = sum(item.product.price * item.quantity for item in cart_items)
        
        if request.method == 'POST':
  
                # Создаем заказ
                order = Order.objects.create(
                    user=request.user,
                    total_price=total_price,

                )
                
                # Создаем элементы заказа. bulk_create Создает множество объектов одним запросом (или несколькими пакетами) в отличе от create, который создает один объект
                OrderItem.objects.bulk_create([
                    OrderItem(
                        order=order,
                        product=item.product,
                        quantity=item.quantity,
                        price=item.product.price
                    ) for item in cart_items
                ])
                
                # Генерируем Excel-чек
                excel_file = generate_excel_receipt(order)
                
                # Отправляем email с чеком
                send_mail(request.user.email, order, excel_file, total_price)
                
                # Очищаем корзину
                cart_items.delete()
                
                messages.success(request, 'Заказ успешно оформлен! Чек отправлен на вашу почту.')
                return redirect('cart_view')
        
        return render(request, 'shop/checkout.html', {
            'cart_items': cart_items,
            'total_price': total_price
        })
        
    except Basket.DoesNotExist:
        messages.warning(request, "Корзина не найдена!")
        return redirect('cart_view')

def generate_excel_receipt(order):
    """Генерирует Excel-чек и возвращает BytesIO объект"""
    wb = Workbook() #создаёт новую книгу Excel (файл).
    ws = wb.active # выбирает активный лист (по умолчанию первый лист).
    ws.title = "Чек заказа" #имя листа в книге.
    
    # Форматирование заголовков
    bold_font = Font(bold=True) # создаёт объект шрифта с жирным начертанием 
    
    # Заголовки таблицы
    headers = ["№", "Товар", "Количество", "Цена", "Сумма"] # названия столбцов
    ws.append(headers) #добавляет первую строку с заголовками в лист Excel.
    for cell in ws[1]:
        cell.font = bold_font # перебираем все ячейки первой строки и сустанавливаем жирный шрифт
    
    # Данные о товарах
    for idx, item in enumerate(order.items.all(), start=1):
        #enumerate(..., start=1) — нумерует товары с 1.
        #order.items.all() — получает все элементы заказа (товары с количеством и ценой).
        ws.append([
            idx,
            item.product.name,
            item.quantity,
            item.price,
            item.price * item.quantity
        ])
    
    # Итоговая сумма
    ws.append(["", "", "", "Итого:", order.total_price])
    for cell in ws[ws.max_row]: #ws.max_row — номер последней строки. и для нее устанавливаем жирный шрифт
        cell.font = bold_font
    
    # Сохраняем в буфер
    buffer = BytesIO() #Создаётся объект BytesIO — это виртуальный файл в памяти.
    wb.save(buffer) #Метод save объекта Workbook сохраняет созданный Excel-файл.
    buffer.seek(0) #seek(0) перемещает указатель на начало файла. потому что при записи он перемещается в конец
    return buffer

def send_mail(email, order, excel_file, total_price):

    email_msg = EmailMessage(
        subject=f'Чек заказа №{order.id} от {order.created.strftime("%d.%m.%Y")}', #subject — тема письма.
        body=f'''
        Благодарим за ваш заказ!
        
        Номер заказа: {order.id}
        Дата: {order.created.strftime("%d.%m.%Y %H:%M")}
        Сумма заказа: {total_price} руб.
        
        Детали заказа в приложенном файле.
        ''',#body — текст письма.
        from_email=settings.DEFAULT_FROM_EMAIL, #from_email — адрес отправителя.берется из настроек джанго
        to=[email], #to — список адресов получателей.
    )
    

    email_msg.attach( #Метод attach() добавляет вложение к письму.
        filename=f'order_{order.id}_receipt.xlsx', #имя файла
        content=excel_file.getvalue(),
        #excel_file.getvalue() — метод объекта BytesIO, который возвращает все данные файла.content — содержимое файла в байтах.
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        #MIME-тип файла, указывающий, что это Excel-файл формата .xlsx. 
    )
    
    email_msg.send()